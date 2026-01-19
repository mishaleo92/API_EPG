"""
Модуль для обробки JSON файлів отриманих від SWOT API
Витягує та структурує статистику з SWOT звітів
"""
import json
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SWOTProcessor:
    """Клас для обробки SWOT JSON файлів"""
    
    def __init__(self):
        """Ініціалізація обробника SWOT файлів"""
        pass
    
    def load_swot_file(self, filepath: str) -> Optional[Dict[str, Any]]:
        """
        Завантажити SWOT JSON файл
        
        Args:
            filepath: Шлях до JSON файлу
            
        Returns:
            Словник з даними або None у разі помилки
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        except FileNotFoundError:
            print(f"Помилка: файл не знайдено: {filepath}")
            return None
        except json.JSONDecodeError as e:
            print(f"Помилка парсингу JSON: {e}")
            return None
        except Exception as e:
            print(f"Помилка завантаження файлу: {e}")
            return None
    
    def extract_statistics(self, swot_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Витягти статистику з SWOT даних
        
        Args:
            swot_data: Дані SWOT звіту
            
        Returns:
            Словник з витягнутою статистикою
        """
        extracted = {
            "kved_statistic": [],
            "intelligence_statistic": [],
            "migration_region_statistic": {},
            "cadastr_estate_statistic": {},
            "status_stats": {},
            "open_close_statistic": {},
            "vehicle_statistic": {}
        }
        
        # Отримуємо дані з різних місць структури
        data = swot_data.get("data") or swot_data.get("raw_response", {}).get("Data") or swot_data.get("raw_response", {}).get("data") or {}
        
        if not data:
            print("Попередження: дані не знайдено в SWOT файлі")
            return extracted
        
        # Рекурсивно шукаємо потрібні поля
        def find_field(obj: Any, target_key: str, path: str = ""):
            """Рекурсивно шукає поле в структурі"""
            if isinstance(obj, dict):
                # Перевіряємо поточний рівень
                if target_key in obj:
                    return obj[target_key]
                
                # Шукаємо в нащадках
                for key, value in obj.items():
                    result = find_field(value, target_key, f"{path}.{key}" if path else key)
                    if result is not None:
                        return result
            
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    result = find_field(item, target_key, f"{path}[{i}]" if path else f"[{i}]")
                    if result is not None:
                        return result
            
            return None
        
        # Витягуємо kvedStatistic
        kved_stat = find_field(data, "kvedStatistic")
        if kved_stat:
            extracted["kved_statistic"] = kved_stat
            print(f"Знайдено {len(kved_stat)} записів КВЕД статистики")
        
        # Витягуємо intelligenceStatistic
        intel_stat = find_field(data, "intelligenceStatistic")
        if intel_stat:
            extracted["intelligence_statistic"] = intel_stat
            print(f"Знайдено {len(intel_stat)} записів статистики інтелекту")
        
        # Витягуємо migrationRegionStatistic (без list, лише загальні цифри)
        migration_stat = find_field(data, "migrationRegionStatistic")
        if migration_stat:
            # Копіюємо лише загальні показники, виключаючи list
            extracted["migration_region_statistic"] = {
                "migrationTotalIn": migration_stat.get("migrationTotalIn"),
                "migrationTotalOut": migration_stat.get("migrationTotalOut"),
                "migrationTotalFopIn": migration_stat.get("migrationTotalFopIn"),
                "migrationTotalFopOut": migration_stat.get("migrationTotalFopOut"),
                "migrationTotalCompanyIn": migration_stat.get("migrationTotalCompanyIn"),
                "migrationTotalCompanyOut": migration_stat.get("migrationTotalCompanyOut")
            }
            print("Знайдено статистику міграції (без списку дат)")
        
        # Витягуємо cadastrEstateStatistic (без cadastrNumbers)
        cadastr_stat = find_field(data, "cadastrEstateStatistic")
        if cadastr_stat:
            # Рекурсивно видаляємо cadastrNumbers з усіх рівнів
            def remove_cadastr_numbers(obj: Any) -> Any:
                """Рекурсивно видаляє cadastrNumbers з об'єкта"""
                if isinstance(obj, dict):
                    # Створюємо новий словник без cadastrNumbers
                    cleaned = {}
                    for key, value in obj.items():
                        if key != "cadastrNumbers":
                            cleaned[key] = remove_cadastr_numbers(value)
                    return cleaned
                elif isinstance(obj, list):
                    # Обробляємо кожен елемент списку
                    return [remove_cadastr_numbers(item) for item in obj]
                else:
                    # Примітивні типи повертаємо як є
                    return obj
            
            extracted["cadastr_estate_statistic"] = remove_cadastr_numbers(cadastr_stat)
            print("Знайдено статистику кадастрових земель (без cadastrNumbers)")
        
        # Витягуємо statusStats
        status_stats = find_field(data, "statusStats")
        if status_stats:
            extracted["status_stats"] = status_stats
            print("Знайдено статистику статусів")
        
        # Витягуємо openCloseStatistic (без list, лише загальні цифри)
        open_close_stat = find_field(data, "openCloseStatistic")
        if open_close_stat:
            # Копіюємо лише загальні показники, виключаючи list
            # У оригінальному SWOT файлі поля без пробілів: companyOpen, fopOpen, тощо
            extracted["open_close_statistic"] = {
                "companyOpen": open_close_stat.get("companyOpen"),
                "companyCurrentClose": open_close_stat.get("companyCurrentClose"),
                "companyPercentLive": open_close_stat.get("companyPercentLive"),
                "fopOpen": open_close_stat.get("fopOpen"),
                "fopCurrentClose": open_close_stat.get("fopCurrentClose"),
                "fopPercentLive": open_close_stat.get("fopPercentLive"),
                "totalOpen": open_close_stat.get("totalOpen"),
                "totalCurrentClose": open_close_stat.get("totalCurrentClose"),
                "totalPercentLive": open_close_stat.get("totalPercentLive")
            }
            # Перевіряємо чи є хоча б одне не-null значення
            has_values = any(v is not None for v in extracted["open_close_statistic"].values())
            if has_values:
                print("Знайдено статистику відкритих/закритих компаній (без списку)")
            else:
                print("Знайдено поле openCloseStatistic, але всі значення null")
        
        # Витягуємо vehicleStatistic
        vehicle_stat = find_field(data, "vehicleStatistic")
        if vehicle_stat:
            extracted["vehicle_statistic"] = {
                "headerCompanyWithCount": vehicle_stat.get("headerCompanyWithCount"),
                "headerCompanyWithoutCount": vehicle_stat.get("headerCompanyWithoutCount"),
                "headerVehicleCount": vehicle_stat.get("headerVehicleCount")
            }
            # Перевіряємо чи є хоча б одне не-null значення
            has_values = any(v is not None for v in extracted["vehicle_statistic"].values())
            if has_values:
                print("Знайдено статистику транспортних засобів")
            else:
                print("Знайдено поле vehicleStatistic, але всі значення null")
        
        return extracted
    
    def save_extracted_statistics(self, statistics: Dict[str, Any], 
                                  input_filepath: str,
                                  output_dir: str = "output") -> Optional[str]:
        """
        Зберегти витягнуту статистику у структурований JSON файл
        
        Args:
            statistics: Витягнута статистика
            input_filepath: Шлях до вхідного SWOT файлу
            output_dir: Директорія для збереження (за замовчуванням "output")
            
        Returns:
            Шлях до збереженого файлу або None у разі помилки
        """
        try:
            # Створюємо директорію якщо вона не існує
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)
            
            # Формуємо ім'я файлу на основі вхідного файлу
            input_file = Path(input_filepath)
            input_stem = input_file.stem
            current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"swot_statistics_{input_stem}_{current_date}.json"
            filepath = output_path / filename
            
            # Створюємо структурований JSON з метаданими
            structured_data = {
                "metadata": {
                    "created_at": datetime.now().isoformat(),
                    "date_created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "source_file": str(input_filepath),
                    "description": f"Витягнута статистика з SWOT звіту. Файл створено {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                },
                "statistics": statistics
            }
            
            # Зберігаємо у JSON файл з красивим форматуванням
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(structured_data, f, ensure_ascii=False, indent=2)
            
            print(f"\nСтатистика збережена у файл: {filepath}")
            return str(filepath)
            
        except Exception as e:
            print(f"Помилка збереження статистики: {e}")
            return None
    
    def save_statistics_to_excel(self, statistics: Dict[str, Any],
                                 input_filepath: str,
                                 output_dir: str = "output") -> Optional[str]:
        """
        Зберегти витягнуту статистику у Excel файл
        
        Args:
            statistics: Витягнута статистика
            input_filepath: Шлях до вхідного SWOT файлу
            output_dir: Директорія для збереження (за замовчуванням "output")
            
        Returns:
            Шлях до збереженого Excel файлу або None у разі помилки
        """
        if not OPENPYXL_AVAILABLE:
            print("Попередження: openpyxl не встановлено. Excel файл не буде створено.")
            print("Встановіть: pip install openpyxl")
            return None
        
        try:
            # Створюємо директорію якщо вона не існує
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)
            
            # Формуємо ім'я файлу
            input_file = Path(input_filepath)
            input_stem = input_file.stem
            current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"swot_statistics_{input_stem}_{current_date}.xlsx"
            filepath = output_path / filename
            
            # Створюємо нову робочу книгу
            wb = Workbook()
            
            # Видаляємо стандартний аркуш
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Стилі для заголовків
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # 1. КВЕД статистика
            if statistics.get("kved_statistic"):
                ws_kved = wb.create_sheet("КВЕД")
                ws_kved.append(["КВЕД ID", "Назва", "Кількість записів"])
                for row in ws_kved.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for item in statistics["kved_statistic"]:
                    ws_kved.append([
                        item.get("kvedId", ""),
                        item.get("name", ""),
                        item.get("qntRecord", 0)
                    ])
                
                # Автопідбір ширини колонок
                for column in ws_kved.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_kved.column_dimensions[column_letter].width = adjusted_width
            
            # 2. Статистика інтелекту (стани)
            if statistics.get("intelligence_statistic"):
                ws_intel = wb.create_sheet("Стани реєстрації")
                ws_intel.append(["Стан", "Кількість записів"])
                for row in ws_intel.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for item in statistics["intelligence_statistic"]:
                    ws_intel.append([
                        item.get("state", ""),
                        item.get("qntRecord", 0)
                    ])
                
                for column in ws_intel.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_intel.column_dimensions[column_letter].width = adjusted_width
            
            # 3. Статистика міграції
            if statistics.get("migration_region_statistic"):
                ws_migration = wb.create_sheet("Міграція")
                migration_stat = statistics["migration_region_statistic"]
                ws_migration.append(["Показник", "Значення"])
                for row in ws_migration.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws_migration.append(["Міграція всього (вхід)", migration_stat.get("migrationTotalIn", 0)])
                ws_migration.append(["Міграція всього (вихід)", migration_stat.get("migrationTotalOut", 0)])
                ws_migration.append(["ФОП міграція (вхід)", migration_stat.get("migrationTotalFopIn", 0)])
                ws_migration.append(["ФОП міграція (вихід)", migration_stat.get("migrationTotalFopOut", 0)])
                ws_migration.append(["Компанії міграція (вхід)", migration_stat.get("migrationTotalCompanyIn", 0)])
                ws_migration.append(["Компанії міграція (вихід)", migration_stat.get("migrationTotalCompanyOut", 0)])
                
                for column in ws_migration.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_migration.column_dimensions[column_letter].width = adjusted_width
            
            # 4. Статистика відкритих/закритих
            open_close_stat = statistics.get("open_close_statistic", {})
            if open_close_stat:
                ws_open_close = wb.create_sheet("Відкриті_Закриті")
                ws_open_close.append(["Показник", "Значення"])
                for row in ws_open_close.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Витягуємо значення, замінюючи None на 0
                # У оригінальному SWOT файлі поля без пробілів
                company_open = open_close_stat.get("companyOpen") if open_close_stat.get("companyOpen") is not None else 0
                company_close = open_close_stat.get("companyCurrentClose") if open_close_stat.get("companyCurrentClose") is not None else 0
                company_percent = open_close_stat.get("companyPercentLive") if open_close_stat.get("companyPercentLive") is not None else 0
                fop_open = open_close_stat.get("fopOpen") if open_close_stat.get("fopOpen") is not None else 0
                fop_close = open_close_stat.get("fopCurrentClose") if open_close_stat.get("fopCurrentClose") is not None else 0
                fop_percent = open_close_stat.get("fopPercentLive") if open_close_stat.get("fopPercentLive") is not None else 0
                total_open = open_close_stat.get("totalOpen") if open_close_stat.get("totalOpen") is not None else 0
                total_close = open_close_stat.get("totalCurrentClose") if open_close_stat.get("totalCurrentClose") is not None else 0
                total_percent = open_close_stat.get("totalPercentLive") if open_close_stat.get("totalPercentLive") is not None else 0
                
                ws_open_close.append(["Компанії відкриті", company_open])
                ws_open_close.append(["Компанії закриті", company_close])
                ws_open_close.append(["Компанії % живих", company_percent])
                ws_open_close.append(["ФОП відкриті", fop_open])
                ws_open_close.append(["ФОП закриті", fop_close])
                ws_open_close.append(["ФОП % живих", fop_percent])
                ws_open_close.append(["Всього відкриті", total_open])
                ws_open_close.append(["Всього закриті", total_close])
                ws_open_close.append(["Всього % живих", total_percent])
                
                for column in ws_open_close.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_open_close.column_dimensions[column_letter].width = adjusted_width
            
            # 5. Статистика транспортних засобів
            vehicle_stat = statistics.get("vehicle_statistic", {})
            if vehicle_stat:
                ws_vehicle = wb.create_sheet("Транспорт")
                ws_vehicle.append(["Показник", "Значення"])
                for row in ws_vehicle.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Витягуємо значення, замінюючи None на 0
                company_with = vehicle_stat.get("headerCompanyWithCount") if vehicle_stat.get("headerCompanyWithCount") is not None else 0
                company_without = vehicle_stat.get("headerCompanyWithoutCount") if vehicle_stat.get("headerCompanyWithoutCount") is not None else 0
                vehicle_count = vehicle_stat.get("headerVehicleCount") if vehicle_stat.get("headerVehicleCount") is not None else 0
                
                ws_vehicle.append(["Компанії з транспортними засобами", company_with])
                ws_vehicle.append(["Компанії без транспортних засобів", company_without])
                ws_vehicle.append(["Всього транспортних засобів", vehicle_count])
                
                for column in ws_vehicle.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_vehicle.column_dimensions[column_letter].width = adjusted_width
            
            # 6. Статистика кадастрових земель (byOwnerForm)
            cadastr_stat = statistics.get("cadastr_estate_statistic", {})
            by_owner = cadastr_stat.get("byOwnerForm", {})
            if by_owner:
                ws_cadastr = wb.create_sheet("Кадастр_Власність")
                ws_cadastr.append(["Тип власності", "Кількість", "Площа", "Ціна НГО"])
                for row in ws_cadastr.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Додаємо загальну статистику
                total = by_owner.get("totalStat", {})
                if total:
                    ws_cadastr.append([
                        total.get("name", "Всього"),
                        total.get("count", 0),
                        total.get("area", 0),
                        total.get("ngoPrice", 0)
                    ])
                
                # Додаємо статистику за типами власності
                stat_list = by_owner.get("statistic", [])
                for item in stat_list:
                    ws_cadastr.append([
                        item.get("name", ""),
                        item.get("count", 0),
                        item.get("area", 0),
                        item.get("ngoPrice", 0) or 0
                    ])
                
                for column in ws_cadastr.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_cadastr.column_dimensions[column_letter].width = adjusted_width
            
            # 6. Статистика кадастрових земель (byPurpose)
            by_purpose = cadastr_stat.get("byPurpose", {})
            if by_purpose:
                ws_purpose = wb.create_sheet("Кадастр_Призначення")
                ws_purpose.append(["Призначення", "Кількість", "Площа", "Ціна НГО"])
                for row in ws_purpose.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Додаємо загальну статистику
                total = by_purpose.get("totalStat", {})
                if total:
                    ws_purpose.append([
                        total.get("name", "Всього"),
                        total.get("count", 0),
                        total.get("area", 0),
                        total.get("ngoPrice", 0)
                    ])
                
                # Додаємо статистику за призначенням
                stat_list = by_purpose.get("statistic", [])
                for item in stat_list:
                    ws_purpose.append([
                        item.get("name", ""),
                        item.get("count", 0),
                        item.get("area", 0),
                        item.get("ngoPrice", 0) or 0
                    ])
                
                for column in ws_purpose.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_purpose.column_dimensions[column_letter].width = adjusted_width
            
            # 7. Статистика статусів (landStat та objectStat)
            status_stats = statistics.get("status_stats", {})
            if status_stats:
                ws_status = wb.create_sheet("Статуси")
                ws_status.append(["Статус", "Земельні ділянки", "Об'єкти"])
                for row in ws_status.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                land_stat = status_stats.get("landStat", {})
                object_stat = status_stats.get("objectStat", {})
                
                # Збираємо всі унікальні ключі
                all_keys = set(land_stat.keys()) | set(object_stat.keys())
                
                for key in sorted(all_keys):
                    ws_status.append([
                        key,
                        land_stat.get(key, 0),
                        object_stat.get(key, 0)
                    ])
                
                for column in ws_status.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_status.column_dimensions[column_letter].width = adjusted_width
            
            # Зберігаємо файл
            wb.save(filepath)
            print(f"Excel файл збережено: {filepath}")
            return str(filepath)
            
        except Exception as e:
            print(f"Помилка створення Excel файлу: {e}")
            return None
    
    def process_swot_file(self, input_filepath: str, 
                         output_dir: str = "output") -> Optional[str]:
        """
        Обробити SWOT файл: завантажити, витягти статистику та зберегти
        
        Args:
            input_filepath: Шлях до SWOT JSON файлу
            output_dir: Директорія для збереження результату
            
        Returns:
            Шлях до збереженого файлу або None у разі помилки
        """
        print(f"\nОбробка SWOT файлу: {input_filepath}")
        print("=" * 60)
        
        # Завантажуємо файл
        swot_data = self.load_swot_file(input_filepath)
        if not swot_data:
            return None
        
        # Витягуємо статистику
        print("\nВитягування статистики...")
        statistics = self.extract_statistics(swot_data)
        
        # Перевіряємо чи щось знайдено
        has_data = any([
            statistics["kved_statistic"],
            statistics["intelligence_statistic"],
            statistics["migration_region_statistic"],
            statistics["cadastr_estate_statistic"],
            statistics["status_stats"],
            statistics["open_close_statistic"],
            statistics["vehicle_statistic"]
        ])
        
        if not has_data:
            print("Попередження: не знайдено жодної статистики у файлі")
        else:
            print(f"\nВитягнуто статистику:")
            print(f"  - КВЕД: {len(statistics['kved_statistic'])} записів")
            print(f"  - Інтелект: {len(statistics['intelligence_statistic'])} записів")
            print(f"  - Міграція: {'Так' if statistics['migration_region_statistic'] else 'Ні'}")
            print(f"  - Кадастр: {'Так' if statistics['cadastr_estate_statistic'] else 'Ні'}")
            print(f"  - Статуси: {'Так' if statistics['status_stats'] else 'Ні'}")
            print(f"  - Відкриті/Закриті: {'Так' if statistics['open_close_statistic'] else 'Ні'}")
            print(f"  - Транспорт: {'Так' if statistics['vehicle_statistic'] else 'Ні'}")
        
        # Зберігаємо результат у JSON
        json_file = self.save_extracted_statistics(statistics, input_filepath, output_dir)
        
        # Зберігаємо результат у Excel
        excel_file = self.save_statistics_to_excel(statistics, input_filepath, output_dir)
        
        return json_file if json_file else excel_file
